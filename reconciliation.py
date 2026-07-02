import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class Reconciler:
    def __init__(self, tolerance: float = 10.0):
        self.tolerance = tolerance
        self.required_columns = [
            "Cod Perm", "Nombre del Estudiante", "Vlr Concepto", "Fecha Emisión", "Nombre Concepto",
            "Nombre Adquiriente", "No Factura", "Nit/Cédula"
        ]

    def _log_error(self, message):
        try:
            with open("log_errores.txt", "a", encoding="utf-8") as f:
                f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")
        except:
            pass

    def load_c3_data(self, file_path: str) -> pd.DataFrame:
        """
        Carga el archivo Excel/CSV de C3 y valida las columnas.
        """
        try:
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path, skiprows=4)
            else:
                df = pd.read_excel(file_path, skiprows=4)
        except Exception as e:
            self._log_error(f"Error leyendo archivo C3: {str(e)}")
            raise ValueError(f"No se pudo leer el archivo. Verifique el formato. Detalle: {str(e)}")

        # Validar columnas
        missing_cols = [col for col in self.required_columns if col not in df.columns]
        if missing_cols:
            error_msg = (f"El archivo C3 no tiene el formato correcto.\n"
                         f"Faltan las siguientes columnas: {', '.join(missing_cols)}\n"
                         f"Columnas encontradas: {', '.join(df.columns)}")
            raise ValueError(error_msg)

        # Mapeo de columnas reales de C3 a la estructura interna
        df = df.rename(columns={
            'Cod Perm': 'DNI_Estudiante',
            'Nombre del Estudiante': 'Nombre_Estudiante',
            'Vlr Concepto': 'Valor_Original',
            'Fecha Emisión': 'Fecha',
            'Nombre Concepto': 'Concepto',
            'Nombre Adquiriente': 'Nombre_Acudiente',
            'No Factura': 'Numero_Factura',
            'Nit/Cédula': 'NIT'
        })

        # Limpiar y estandarizar datos
        df['DNI_Estudiante'] = df['DNI_Estudiante'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        df['Valor_Original'] = pd.to_numeric(df['Valor_Original'], errors='coerce').fillna(0)
        
        # Calcular Débito (Cobrado) y Crédito (Recibido)
        # Débito = positivo, Crédito = negativo (se convierte a absoluto)
        # Nota: 'Vlr Concepto' se mapea a 'Valor_Original'. El valor recibido se
        # pasa como absoluto a 'Valor_Recibido', el cual es consumido por la
        # facturación masiva para llamar a create_receipt con el precio/valor del pago.
        # Esto permite que el cliente de Siigo reciba el valor correcto en el campo
        # 'value' del arreglo de payments.
        df['Valor_Cobrado'] = df['Valor_Original'].apply(lambda x: x if x > 0 else 0)
        df['Valor_Recibido'] = df['Valor_Original'].apply(lambda x: abs(x) if x < 0 else 0)

        # Keep the C3 rows disaggregated (one row per original entry). Do NOT group.
        # Ensure columns exist for downstream code
        df = df.reset_index(drop=True)
        df['Saldo_Pendiente'] = df['Valor_Cobrado'] - df['Valor_Recibido']
        # Compatibility alias used elsewhere
        df['Valor_Facturado'] = df['Valor_Cobrado']
        # Create an idempotency key per row: prefer Numero_Factura if present, otherwise build a deterministic key
        def make_key(row, idx):
            num = str(row.get('Numero_Factura', '')).strip()
            dni = str(row.get('DNI_Estudiante', '')).strip()
            if num and num.lower() != 'nan':
                return f"c3-{num}"
            # deterministic fallback: c3-{dni}-{index}
            return f"c3-{dni}-{idx}"

        df['Idempotency_Key'] = [make_key(r, i) for i, r in df.reset_index(drop=True).iterrows()]
        return df

    def reconcile(self, df_siigo: pd.DataFrame, df_c3: pd.DataFrame) -> pd.DataFrame:
        """
        Realiza el Outer Join. Para C3 (múltiples conceptos posibles), cruzamos con Siigo.
        """
        # New matching strategy: perform one-to-one matching between disaggregated C3 rows
        # and Siigo individual invoices. Avoid cartesian product by iterating C3 rows
        # and trying to find the best match in Siigo (by DNI + exact amount or DNI + Concept).

        # Prepare copies to mark matched Siigo rows
        siigo = df_siigo.copy()
        siigo['__matched'] = False

        results = []

        # Helper to try find a siigo index by criteria
        def find_siigo_match(dni, amount, concept=None):
            # First try exact DNI + amount
            candidates = siigo[(siigo['DNI_Estudiante'] == dni) & (~siigo['__matched'])]
            if not candidates.empty:
                # Try exact amount match
                exact = candidates[candidates['Valor_Facturado'].fillna(0).astype(float) == float(amount)]
                if not exact.empty:
                    return exact.index[0]
                # Try match by concept string contained in Items
                if concept is not None:
                    by_concept = candidates[candidates['Items'].fillna('').str.contains(str(concept), case=False, na=False)]
                    if not by_concept.empty:
                        return by_concept.index[0]
                # Fallback: if only one candidate for the DNI, take it
                if len(candidates) == 1:
                    return candidates.index[0]
            return None

        for idx, c3_row in df_c3.reset_index(drop=True).iterrows():
            dni = str(c3_row.get('DNI_Estudiante', '')).strip()
            concepto = c3_row.get('Concepto', '')
            valor_cobrado = float(c3_row.get('Valor_Cobrado', 0) or 0)

            match_idx = find_siigo_match(dni, valor_cobrado, concepto)
            matched_siigo = None

            if match_idx is not None:
                matched_siigo = siigo.loc[match_idx]
                siigo.at[match_idx, '__matched'] = True

            if matched_siigo is None:
                estado = "ANOMALÍA TIPO A (No en Siigo)"
                siigo_vals = {k: np.nan for k in ['Invoice_Id','Invoice_Number','Valor_Facturado','Fecha','Items','Nombre_Estudiante']}
            else:
                valor_siigo = float(matched_siigo.get('Valor_Facturado', 0) or 0)
                diff = abs(valor_cobrado - valor_siigo)
                if diff <= self.tolerance:
                    estado = 'MATCH'
                else:
                    estado = f'ANOMALÍA TIPO B (Dif: {diff:.2f})'
                siigo_vals = matched_siigo.to_dict()

            result_row = {
                'DNI_Estudiante': dni,
                'Estado': estado,
                'Nombre_Estudiante_C3': c3_row.get('Nombre_Estudiante'),
                'Nombre_Acudiente': c3_row.get('Nombre_Acudiente'),
                'NIT': c3_row.get('NIT'),
                'Numero_Factura': c3_row.get('Numero_Factura'),
                'Concepto': concepto,
                'Valor_Cobrado': valor_cobrado,
                'Valor_Recibido': float(c3_row.get('Valor_Recibido', 0) or 0),
                'Saldo_Pendiente': c3_row.get('Saldo_Pendiente'),
                'Valor_Facturado_Siigo': siigo_vals.get('Valor_Facturado') if siigo_vals is not None else np.nan,
                'Nombre_Estudiante_Siigo': siigo_vals.get('Nombre_Estudiante') if siigo_vals is not None else np.nan,
                'Invoice_Id': siigo_vals.get('Invoice_Id') if siigo_vals is not None else np.nan,
                'Invoice_Number': siigo_vals.get('Invoice_Number') if siigo_vals is not None else np.nan,
                'Items_Siigo': siigo_vals.get('Items') if siigo_vals is not None else np.nan,
                'Fecha_C3': c3_row.get('Fecha'),
                'Fecha_Siigo': siigo_vals.get('Fecha') if siigo_vals is not None else np.nan
            }
            results.append(result_row)

        # Any remaining unmatched Siigo rows -> Anomalía Tipo C (No en C3)
        unmatched_siigo = siigo[~siigo['__matched']]
        for _, srow in unmatched_siigo.iterrows():
            results.append({
                'DNI_Estudiante': srow.get('DNI_Estudiante'),
                'Estado': 'ANOMALÍA TIPO C (No en C3)',
                'Nombre_Estudiante_C3': np.nan,
                'Nombre_Acudiente': np.nan,
                'NIT': np.nan,
                'Numero_Factura': np.nan,
                'Concepto': np.nan,
                'Valor_Cobrado': np.nan,
                'Valor_Recibido': np.nan,
                'Saldo_Pendiente': np.nan,
                'Valor_Facturado_Siigo': srow.get('Valor_Facturado'),
                'Nombre_Estudiante_Siigo': srow.get('Nombre_Estudiante'),
                'Invoice_Id': srow.get('Invoice_Id'),
                'Invoice_Number': srow.get('Invoice_Number'),
                'Items_Siigo': srow.get('Items'),
                'Fecha_C3': np.nan,
                'Fecha_Siigo': srow.get('Fecha')
            })

        result_df = pd.DataFrame(results)
        return result_df

    def export_to_excel(self, df_result: pd.DataFrame, output_path: str):
        """
        Exporta el resultado a Excel y aplica formato condicional.
        """
        df_result.to_excel(output_path, index=False, engine='openpyxl')
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        estado_col_idx = None
        for cell in ws[1]:
            if cell.value == "Estado":
                estado_col_idx = cell.column
                break
                
        if estado_col_idx is None:
            wb.save(output_path)
            return
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            estado_cell = ws.cell(row=row[0].row, column=estado_col_idx)
            estado_value = str(estado_cell.value).upper()
            
            if "ANOMALÍA" in estado_value or "ERROR" in estado_value:
                for cell in row:
                    cell.fill = red_fill
                    
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 40)
            
        wb.save(output_path)
